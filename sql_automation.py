"""
sql_automation.py
-----------------
Automated SQL pipeline for Telco Customer Churn Analysis.
Connects to MySQL, loads engineered features, runs 14 analytical
queries and exports results to a formatted Excel report.

Author  : Sabari A S
LinkedIn: linkedin.com/in/sabari3299
"""

import time
import pandas as pd
import numpy as np
import mysql.connector
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import warnings
warnings.filterwarnings("ignore")

# ── Config ───────────────────────────────────────────────
MYSQL_HOST     = "localhost"
MYSQL_USER     = "root"
MYSQL_PASSWORD = "1234"
MYSQL_DB       = "churn_db"
DATA_PATH      = r"E:\Customer churn\Data\customer churn.csv"
OUTPUT_DIR     = r"E:\Customer churn\sql_results"

os.makedirs(OUTPUT_DIR, exist_ok=True)

print("=" * 55)
print("  AUTOMATED SQL PIPELINE")
print("=" * 55)

# ── Connect to MySQL ─────────────────────────────────────
print("\n📡 Connecting to MySQL...")
conn = mysql.connector.connect(
    host=MYSQL_HOST,
    user=MYSQL_USER,
    password=MYSQL_PASSWORD
)
cursor = conn.cursor()
cursor.execute(f"CREATE DATABASE IF NOT EXISTS {MYSQL_DB}")
cursor.execute(f"USE {MYSQL_DB}")
conn.commit()
print(f"✅ Database '{MYSQL_DB}' ready")

# ── Load & Engineer Features ─────────────────────────────
print("\n📂 Loading CSV into MySQL...")
df = pd.read_csv(DATA_PATH)
df["TotalCharges"]          = pd.to_numeric(df["TotalCharges"], errors="coerce")
df["TotalCharges"]          = df["TotalCharges"].fillna(df["TotalCharges"].median())
df["Churn_Flag"]            = (df["Churn"] == "Yes").astype(int)
df["TenureBand"]            = pd.cut(
    df["tenure"],
    bins=[0, 12, 24, 48, 72],
    labels=["0-12 Months", "13-24 Months", "25-48 Months", "49-72 Months"]
)
df["RevenueTier"]           = pd.cut(
    df["MonthlyCharges"],
    bins=[0, 35, 70, 120],
    labels=["Low", "Medium", "High"]
)
df["CustomerLifetimeValue"] = df["MonthlyCharges"] * df["tenure"]
df["RevenueRisk"]           = np.where(df["Churn"] == "Yes", df["MonthlyCharges"], 0)
df["RiskScore"]             = (
    np.where(df["Contract"]        == "Month-to-month", 40, 0)
    + np.where(df["InternetService"] == "Fiber optic",   30, 0)
    + np.where(df["tenure"]          <= 12,              30, 0)
)
df["RiskCategory"]          = pd.cut(
    df["RiskScore"],
    bins=[-1, 30, 60, 100],
    labels=["Low", "Medium", "High"]
)

# Build summary BEFORE renaming columns
summary = pd.DataFrame({
    "Metric": ["Total Customers", "Churn Rate (%)", "Revenue At Risk ($/mo)"],
    "Value": [
        len(df),
        round(df["Churn_Flag"].mean() * 100, 2),
        round(df["RevenueRisk"].sum(), 2)
    ]
})

df.columns = df.columns.str.replace(" ", "_").str.lower()

engine = create_engine(
    f"mysql+mysqlconnector://{MYSQL_USER}:{MYSQL_PASSWORD}@{MYSQL_HOST}/{MYSQL_DB}"
)
df.to_sql("telco_churn", con=engine, if_exists="replace", index=False)
print(f"✅ {len(df):,} rows loaded into MySQL")

# ── Queries ──────────────────────────────────────────────
queries = {
    "Q1_overall_churn": """
        SELECT COUNT(*) AS total,
               SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END) AS churned,
               ROUND(SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END)/COUNT(*)*100,2) AS churn_rate_pct
        FROM telco_churn
    """,
    "Q2_by_contract": """
        SELECT contract,
               COUNT(*) AS total,
               SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END) AS churned,
               ROUND(SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END)/COUNT(*)*100,2) AS churn_rate_pct
        FROM telco_churn
        GROUP BY contract
        ORDER BY churn_rate_pct DESC
    """,
    "Q3_by_internet": """
        SELECT internetservice,
               COUNT(*) AS total,
               SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END) AS churned,
               ROUND(SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END)/COUNT(*)*100,2) AS churn_rate_pct
        FROM telco_churn
        GROUP BY internetservice
        ORDER BY churn_rate_pct DESC
    """,
    "Q4_by_payment": """
        SELECT paymentmethod,
               COUNT(*) AS total,
               SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END) AS churned,
               ROUND(SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END)/COUNT(*)*100,2) AS churn_rate_pct
        FROM telco_churn
        GROUP BY paymentmethod
        ORDER BY churn_rate_pct DESC
    """,
    "Q5_tenure_charges": """
        SELECT churn,
               ROUND(AVG(tenure),1)         AS avg_tenure,
               ROUND(AVG(monthlycharges),2) AS avg_monthly_charges
        FROM telco_churn
        GROUP BY churn
    """,
    "Q6_revenue_at_risk": """
        SELECT ROUND(SUM(monthlycharges),2) AS revenue_at_risk,
               ROUND(SUM(monthlycharges)/(SELECT SUM(monthlycharges) FROM telco_churn)*100,2) AS pct_of_total
        FROM telco_churn
        WHERE churn='Yes'
    """,
    "Q7_senior_churn": """
        SELECT CASE WHEN seniorcitizen=1 THEN 'Senior' ELSE 'Non-Senior' END AS segment,
               COUNT(*) AS total,
               ROUND(SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END)/COUNT(*)*100,2) AS churn_rate_pct
        FROM telco_churn
        GROUP BY seniorcitizen
    """,
    "Q8_tenure_cohort": """
        SELECT CASE
                 WHEN tenure BETWEEN 0  AND 12 THEN '0-12 months'
                 WHEN tenure BETWEEN 13 AND 24 THEN '13-24 months'
                 WHEN tenure BETWEEN 25 AND 48 THEN '25-48 months'
                 ELSE '49+ months'
               END AS tenure_bucket,
               COUNT(*) AS total,
               ROUND(SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END)/COUNT(*)*100,2) AS churn_rate_pct
        FROM telco_churn
        GROUP BY tenure_bucket
        ORDER BY MIN(tenure)
    """,
    "Q9_high_risk_segment": """
        SELECT customerid, contract, internetservice, paymentmethod,
               tenure, monthlycharges, churn
        FROM telco_churn
        WHERE contract='Month-to-month'
          AND tenure <= 12
          AND internetservice='Fiber optic'
          AND churn='Yes'
        ORDER BY monthlycharges DESC
        LIMIT 20
    """,
    "Q10_window_rank": """
        SELECT customerid, contract, monthlycharges, churn,
               RANK() OVER (PARTITION BY contract ORDER BY monthlycharges DESC) AS charge_rank,
               ROUND(AVG(monthlycharges) OVER (PARTITION BY contract),2)        AS avg_in_contract
        FROM telco_churn
        ORDER BY contract, charge_rank
        LIMIT 30
    """,
    "Q11_crosstab": """
        SELECT contract,
               ROUND(AVG(CASE WHEN internetservice='Fiber optic'
                              THEN CASE WHEN churn='Yes' THEN 1.0 ELSE 0 END END)*100,1) AS fiber_churn_pct,
               ROUND(AVG(CASE WHEN internetservice='DSL'
                              THEN CASE WHEN churn='Yes' THEN 1.0 ELSE 0 END END)*100,1) AS dsl_churn_pct
        FROM telco_churn
        GROUP BY contract
        ORDER BY contract
    """,
    "Q12_cumulative_churn": """
        SELECT tenure,
               COUNT(*) AS customers,
               SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END) AS churned,
               SUM(SUM(CASE WHEN churn='Yes' THEN 1 ELSE 0 END))
                   OVER (ORDER BY tenure ROWS UNBOUNDED PRECEDING) AS cumulative_churn
        FROM telco_churn
        GROUP BY tenure
        ORDER BY tenure
    """,
    "Q13_risk_category": """
        SELECT riskcategory,
               COUNT(*)                          AS customers,
               ROUND(AVG(churn_flag) * 100, 2)  AS churn_rate_pct
        FROM telco_churn
        GROUP BY riskcategory
        ORDER BY churn_rate_pct DESC
    """,
    "Q14_revenue_by_tenure": """
        SELECT churn,
               tenureband,
               COUNT(*)                          AS customers,
               ROUND(SUM(monthlycharges), 2)     AS total_monthly_revenue,
               ROUND(AVG(monthlycharges), 2)     AS avg_monthly_charges
        FROM telco_churn
        GROUP BY churn, tenureband
        ORDER BY tenureband, churn
    """,
}

# ── Run Queries & Export ─────────────────────────────────
print("\n🔍 Running all 14 queries...")
print("-" * 55)

excel_path = os.path.join(OUTPUT_DIR, "churn_sql_results.xlsx")

with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
    for name, query in queries.items():
        try:
            t0     = time.time()
            result = pd.read_sql(query, engine)
            elapsed = round(time.time() - t0, 2)
            result.to_excel(writer, sheet_name=name[:31], index=False)
            print(f"  ✅ {name} → {len(result)} rows ({elapsed}s)")
        except Exception as e:
            print(f"  ❌ {name} failed: {e}")

    summary.to_excel(writer, sheet_name="Executive_Summary", index=False)

# ── Format Excel ─────────────────────────────────────────
wb = load_workbook(excel_path)
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="378ADD")
center      = Alignment(horizontal="center")

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

wb.save(excel_path)

print("-" * 55)
print(f"\n✅ Results saved → {excel_path}")
print("✅ Excel formatting applied")
print("\n" + "=" * 55)
print("  SQL PIPELINE COMPLETE")
print("=" * 55)
