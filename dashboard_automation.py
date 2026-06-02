"""
============================================================
 Customer Churn Analysis — Automated Dashboard Generator
 Author  : Sabari A S
 LinkedIn: linkedin.com/in/sabari3299
============================================================
 This script automatically generates:
 1. All dashboard visuals as PNG files
 2. A full Excel report with all metrics
 3. An automated PowerPoint presentation
 4. A summary HTML dashboard
============================================================
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import seaborn as sns
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils import get_column_letter
import os
import warnings
warnings.filterwarnings("ignore")

# ── Config ───────────────────────────────────────────────
DATA_PATH   = r"E:\Customer churn\Data\customer churn.csv"
OUTPUT_DIR  = r"E:\Customer churn\dashboard_output"
VISUALS_DIR = r"E:\Customer churn\visuals"

os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(VISUALS_DIR, exist_ok=True)

# ── Colors ───────────────────────────────────────────────
CHURN_COLOR   = "#E24B4A"
RETAIN_COLOR  = "#1D9E75"
NEUTRAL_COLOR = "#378ADD"
AMBER_COLOR   = "#FAC775"

print("="*60)
print("  AUTOMATED DASHBOARD GENERATOR")
print("="*60)

# ══════════════════════════════════════════════════════════
# STEP 1 — Load & Prepare Data
# ══════════════════════════════════════════════════════════
print("\n📂 Loading data...")
df = pd.read_csv(DATA_PATH)
df["TotalCharges"] = pd.to_numeric(df["TotalCharges"], errors="coerce")
df["TotalCharges"].fillna(df["TotalCharges"].median(), inplace=True)
df["Churn_Flag"] = (df["Churn"] == "Yes").astype(int)
print(f"✅ {len(df):,} rows loaded")

# ══════════════════════════════════════════════════════════
# STEP 2 — Calculate All KPI Measures
# ══════════════════════════════════════════════════════════
print("\n📊 Calculating KPI measures...")

measures = {
    "Total Customers"       : len(df),
    "Churned Customers"     : df["Churn_Flag"].sum(),
    "Retained Customers"    : (df["Churn_Flag"] == 0).sum(),
    "Churn Rate"            : f"{df['Churn_Flag'].mean():.1%}",
    "Revenue at Risk"       : f"${df[df['Churn']=='Yes']['MonthlyCharges'].sum():,.0f}/mo",
    "Total Monthly Revenue" : f"${df['MonthlyCharges'].sum():,.0f}/mo",
    "Avg Monthly Charges"   : f"${df['MonthlyCharges'].mean():.2f}",
    "Avg Tenure Churned"    : f"{df[df['Churn']=='Yes']['tenure'].mean():.1f} months",
    "Avg Tenure Retained"   : f"{df[df['Churn']=='No']['tenure'].mean():.1f} months",
    "Revenue Retention Rate": f"{(1 - df[df['Churn']=='Yes']['MonthlyCharges'].sum()/df['MonthlyCharges'].sum()):.1%}",
    "High Risk Segment Size": len(df[(df['Contract']=='Month-to-month') & (df['tenure']<=12) & (df['InternetService']=='Fiber optic')]),
}

for k, v in measures.items():
    print(f"   {k:<30} → {v}")

# ══════════════════════════════════════════════════════════
# STEP 3 — Generate All Visuals
# ══════════════════════════════════════════════════════════
print("\n🎨 Generating visuals...")

sns.set_theme(style="whitegrid")
plt.rcParams.update({
    "figure.dpi"        : 150,
    "figure.facecolor"  : "white",
    "axes.facecolor"    : "#FAFAFA",
    "axes.spines.top"   : False,
    "axes.spines.right" : False,
})

# Visual 1 — Churn Distribution
fig, ax = plt.subplots(figsize=(6, 4))
counts = df["Churn"].value_counts()
bars = ax.bar(counts.index, counts.values,
              color=[RETAIN_COLOR, CHURN_COLOR], width=0.5, edgecolor="white")
for bar, val in zip(bars, counts.values):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 30,
            f"{val:,}\n({val/len(df):.1%})", ha="center", fontsize=11, fontweight="bold")
ax.set_title("Overall Churn Distribution", fontsize=14, fontweight="bold", pad=15)
ax.set_ylabel("Number of Customers")
ax.set_ylim(0, counts.max() * 1.2)
plt.tight_layout()
plt.savefig(f"{VISUALS_DIR}/01_churn_distribution.png", bbox_inches="tight")
plt.close()
print("   ✅ 01_churn_distribution.png")

# Visual 2 — Churn by Contract
fig, ax = plt.subplots(figsize=(7, 4))
data = df.groupby("Contract")["Churn_Flag"].mean().sort_values(ascending=False) * 100
bars = ax.barh(data.index, data.values,
               color=[CHURN_COLOR, AMBER_COLOR, RETAIN_COLOR], edgecolor="white")
for bar, val in zip(bars, data.values):
    ax.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height()/2,
            f"{val:.1f}%", va="center", fontsize=11, fontweight="bold")
ax.set_title("Churn Rate by Contract Type", fontsize=14, fontweight="bold", pad=15)
ax.set_xlabel("Churn Rate (%)")
ax.set_xlim(0, data.max() * 1.3)
plt.tight_layout()
plt.savefig(f"{VISUALS_DIR}/02_churn_by_contract.png", bbox_inches="tight")
plt.close()
print("   ✅ 02_churn_by_contract.png")

# Visual 3 — Tenure Distribution
fig, ax = plt.subplots(figsize=(8, 4))
df[df["Churn"]=="No"]["tenure"].hist(ax=ax, bins=30, alpha=0.7,
    color=RETAIN_COLOR, label="Retained", edgecolor="white")
df[df["Churn"]=="Yes"]["tenure"].hist(ax=ax, bins=30, alpha=0.7,
    color=CHURN_COLOR, label="Churned", edgecolor="white")
ax.set_title("Tenure Distribution — Churned vs Retained",
             fontsize=14, fontweight="bold", pad=15)
ax.set_xlabel("Tenure (months)")
ax.set_ylabel("Number of Customers")
ax.legend(fontsize=11)
plt.tight_layout()
plt.savefig(f"{VISUALS_DIR}/03_tenure_distribution.png", bbox_inches="tight")
plt.close()
print("   ✅ 03_tenure_distribution.png")

# Visual 4 — Monthly Charges Boxplot
fig, ax = plt.subplots(figsize=(6, 4))
df.boxplot(column="MonthlyCharges", by="Churn", ax=ax, patch_artist=True,
           boxprops=dict(facecolor="#B5D4F4", color=NEUTRAL_COLOR),
           medianprops=dict(color=CHURN_COLOR, linewidth=2))
ax.set_title("Monthly Charges — Churned vs Retained",
             fontsize=14, fontweight="bold")
ax.set_xlabel("Churn")
ax.set_ylabel("Monthly Charges ($)")
plt.suptitle("")
plt.tight_layout()
plt.savefig(f"{VISUALS_DIR}/04_monthly_charges.png", bbox_inches="tight")
plt.close()
print("   ✅ 04_monthly_charges.png")

# Visual 5 — Internet Service
fig, ax = plt.subplots(figsize=(7, 4))
data = df.groupby("InternetService")["Churn_Flag"].mean().sort_values(ascending=False) * 100
bars = ax.bar(data.index, data.values,
              color=[CHURN_COLOR, AMBER_COLOR, RETAIN_COLOR],
              width=0.5, edgecolor="white")
for bar, val in zip(bars, data.values):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5,
            f"{val:.1f}%", ha="center", fontsize=11, fontweight="bold")
ax.set_title("Churn Rate by Internet Service", fontsize=14, fontweight="bold", pad=15)
ax.set_ylabel("Churn Rate (%)")
ax.set_ylim(0, data.max() * 1.25)
plt.tight_layout()
plt.savefig(f"{VISUALS_DIR}/05_churn_by_internet.png", bbox_inches="tight")
plt.close()
print("   ✅ 05_churn_by_internet.png")

# Visual 6 — Payment Method
fig, ax = plt.subplots(figsize=(8, 4))
data = df.groupby("PaymentMethod")["Churn_Flag"].mean().sort_values(ascending=False) * 100
short = [p.replace(" (automatic)", "\n(auto)") for p in data.index]
bars = ax.barh(short, data.values,
               color=[CHURN_COLOR, AMBER_COLOR, RETAIN_COLOR, RETAIN_COLOR],
               edgecolor="white")
for bar, val in zip(bars, data.values):
    ax.text(bar.get_width() + 0.3, bar.get_y() + bar.get_height()/2,
            f"{val:.1f}%", va="center", fontsize=10, fontweight="bold")
ax.set_title("Churn Rate by Payment Method", fontsize=14, fontweight="bold", pad=15)
ax.set_xlabel("Churn Rate (%)")
ax.set_xlim(0, data.max() * 1.3)
plt.tight_layout()
plt.savefig(f"{VISUALS_DIR}/06_churn_by_payment.png", bbox_inches="tight")
plt.close()
print("   ✅ 06_churn_by_payment.png")

# Visual 7 — Heatmap
fig, ax = plt.subplots(figsize=(8, 4))
pivot = df.pivot_table(values="Churn_Flag", index="Contract",
                       columns="InternetService", aggfunc="mean") * 100
sns.heatmap(pivot, annot=True, fmt=".1f", cmap="RdYlGn_r",
            linewidths=0.5, ax=ax, cbar_kws={"label": "Churn Rate (%)"})
ax.set_title("Churn Rate (%) — Contract × Internet Service",
             fontsize=14, fontweight="bold", pad=15)
plt.tight_layout()
plt.savefig(f"{VISUALS_DIR}/07_churn_heatmap.png", bbox_inches="tight")
plt.close()
print("   ✅ 07_churn_heatmap.png")

# Visual 8 — KPI Summary
fig, axes = plt.subplots(2, 4, figsize=(14, 6))
fig.suptitle("Customer Churn Analysis — KPI Summary",
             fontsize=16, fontweight="bold", y=1.02)

kpis = [
    ("Total Customers",    f"{len(df):,}",                                    NEUTRAL_COLOR),
    ("Churned Customers",  f"{df['Churn_Flag'].sum():,}",                      CHURN_COLOR),
    ("Churn Rate",         f"{df['Churn_Flag'].mean():.1%}",                   CHURN_COLOR),
    ("Revenue at Risk",    f"${df[df['Churn']=='Yes']['MonthlyCharges'].sum():,.0f}",CHURN_COLOR),
    ("Retained Customers", f"{(df['Churn_Flag']==0).sum():,}",                 RETAIN_COLOR),
    ("Monthly Revenue",    f"${df['MonthlyCharges'].sum():,.0f}",               RETAIN_COLOR),
    ("Avg Tenure Churned", f"{df[df['Churn']=='Yes']['tenure'].mean():.1f} mo", NEUTRAL_COLOR),
    ("Avg Tenure Retained",f"{df[df['Churn']=='No']['tenure'].mean():.1f} mo",  RETAIN_COLOR),
]

for ax, (label, value, color) in zip(axes.flat, kpis):
    ax.text(0.5, 0.6, value, transform=ax.transAxes, fontsize=22,
            fontweight="bold", ha="center", va="center", color=color)
    ax.text(0.5, 0.25, label, transform=ax.transAxes, fontsize=10,
            ha="center", va="center", color="#444")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis("off")
    ax.set_facecolor("#F8F8F8")
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_edgecolor("#DDD")

plt.tight_layout()
plt.savefig(f"{VISUALS_DIR}/08_kpi_summary.png", bbox_inches="tight")
plt.close()
print("   ✅ 08_kpi_summary.png")

# ══════════════════════════════════════════════════════════
# STEP 4 — Generate Excel Report
# ══════════════════════════════════════════════════════════
print("\n📊 Generating Excel report...")

wb = openpyxl.Workbook()

# ── Sheet 1: KPI Summary ─────────────────────────────────
ws1 = wb.active
ws1.title = "KPI Summary"

header_fill   = PatternFill("solid", fgColor="1D9E75")
churn_fill    = PatternFill("solid", fgColor="E24B4A")
subhead_fill  = PatternFill("solid", fgColor="378ADD")
white_font    = Font(color="FFFFFF", bold=True, size=12)
bold_font     = Font(bold=True, size=11)
center        = Alignment(horizontal="center", vertical="center")

ws1.merge_cells("A1:C1")
ws1["A1"] = "CUSTOMER CHURN ANALYSIS — KPI SUMMARY"
ws1["A1"].font = Font(color="FFFFFF", bold=True, size=14)
ws1["A1"].fill = header_fill
ws1["A1"].alignment = center
ws1.row_dimensions[1].height = 35

ws1.merge_cells("A2:C2")
ws1["A2"] = "Built by Sabari A S | linkedin.com/in/sabari3299"
ws1["A2"].font = Font(color="FFFFFF", size=10)
ws1["A2"].fill = subhead_fill
ws1["A2"].alignment = center

ws1["A4"] = "Metric"
ws1["B4"] = "Value"
ws1["C4"] = "Status"
for cell in [ws1["A4"], ws1["B4"], ws1["C4"]]:
    cell.font = white_font
    cell.fill = subhead_fill
    cell.alignment = center

rows = [
    ("Total Customers",        len(df),                                          "✅"),
    ("Churned Customers",      int(df["Churn_Flag"].sum()),                      "🔴"),
    ("Retained Customers",     int((df["Churn_Flag"]==0).sum()),                 "✅"),
    ("Churn Rate",             f"{df['Churn_Flag'].mean():.1%}",                 "🔴"),
    ("Revenue at Risk/Month",  f"${df[df['Churn']=='Yes']['MonthlyCharges'].sum():,.0f}", "🔴"),
    ("Total Monthly Revenue",  f"${df['MonthlyCharges'].sum():,.0f}",            "✅"),
    ("Avg Monthly Charges",    f"${df['MonthlyCharges'].mean():.2f}",            "📊"),
    ("Avg Tenure Churned",     f"{df[df['Churn']=='Yes']['tenure'].mean():.1f} months", "🔴"),
    ("Avg Tenure Retained",    f"{df[df['Churn']=='No']['tenure'].mean():.1f} months",  "✅"),
    ("Revenue Retention Rate", f"{(1-df[df['Churn']=='Yes']['MonthlyCharges'].sum()/df['MonthlyCharges'].sum()):.1%}", "✅"),
]

for i, (metric, value, status) in enumerate(rows, start=5):
    ws1[f"A{i}"] = metric
    ws1[f"B{i}"] = value
    ws1[f"C{i}"] = status
    ws1[f"A{i}"].font = bold_font
    ws1[f"B{i}"].alignment = center
    ws1[f"C{i}"].alignment = center
    if i % 2 == 0:
        for col in ["A", "B", "C"]:
            ws1[f"{col}{i}"].fill = PatternFill("solid", fgColor="F0F0F0")

ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 20
ws1.column_dimensions["C"].width = 10

# ── Sheet 2: Churn by Contract ───────────────────────────
ws2 = wb.create_sheet("By Contract")
contract_data = df.groupby("Contract").agg(
    Total=("Churn_Flag", "count"),
    Churned=("Churn_Flag", "sum")
).reset_index()
contract_data["Churn Rate"] = (contract_data["Churned"] / contract_data["Total"] * 100).round(2)

ws2["A1"] = "Contract"
ws2["B1"] = "Total"
ws2["C1"] = "Churned"
ws2["D1"] = "Churn Rate %"
for cell in [ws2["A1"], ws2["B1"], ws2["C1"], ws2["D1"]]:
    cell.font = white_font
    cell.fill = subhead_fill
    cell.alignment = center

for i, row in contract_data.iterrows():
    ws2[f"A{i+2}"] = row["Contract"]
    ws2[f"B{i+2}"] = row["Total"]
    ws2[f"C{i+2}"] = row["Churned"]
    ws2[f"D{i+2}"] = row["Churn Rate"]

for col in ["A", "B", "C", "D"]:
    ws2.column_dimensions[col].width = 18

# ── Sheet 3: Churn by Internet ───────────────────────────
ws3 = wb.create_sheet("By Internet Service")
inet_data = df.groupby("InternetService").agg(
    Total=("Churn_Flag", "count"),
    Churned=("Churn_Flag", "sum")
).reset_index()
inet_data["Churn Rate"] = (inet_data["Churned"] / inet_data["Total"] * 100).round(2)

ws3["A1"] = "Internet Service"
ws3["B1"] = "Total"
ws3["C1"] = "Churned"
ws3["D1"] = "Churn Rate %"
for cell in [ws3["A1"], ws3["B1"], ws3["C1"], ws3["D1"]]:
    cell.font = white_font
    cell.fill = subhead_fill
    cell.alignment = center

for i, row in inet_data.iterrows():
    ws3[f"A{i+2}"] = row["InternetService"]
    ws3[f"B{i+2}"] = row["Total"]
    ws3[f"C{i+2}"] = row["Churned"]
    ws3[f"D{i+2}"] = row["Churn Rate"]

for col in ["A", "B", "C", "D"]:
    ws3.column_dimensions[col].width = 20

# ── Sheet 4: High Risk Segment ───────────────────────────
ws4 = wb.create_sheet("High Risk Segment")
high_risk = df[
    (df["Contract"] == "Month-to-month") &
    (df["tenure"] <= 12) &
    (df["InternetService"] == "Fiber optic") &
    (df["Churn"] == "Yes")
][["customerID", "Contract", "InternetService",
   "PaymentMethod", "tenure", "MonthlyCharges", "Churn"]]

cols = ["customerID", "Contract", "InternetService",
        "PaymentMethod", "tenure", "MonthlyCharges", "Churn"]
for j, col in enumerate(cols, start=1):
    cell = ws4.cell(row=1, column=j, value=col)
    cell.font = white_font
    cell.fill = churn_fill
    cell.alignment = center

for i, row in high_risk.iterrows():
    for j, col in enumerate(cols, start=1):
        ws4.cell(row=i+2, column=j, value=row[col])

for col_idx in range(1, len(cols)+1):
    ws4.column_dimensions[get_column_letter(col_idx)].width = 20

excel_path = os.path.join(OUTPUT_DIR, "churn_dashboard_report.xlsx")
wb.save(excel_path)
print(f"✅ Excel report saved → {excel_path}")

# ══════════════════════════════════════════════════════════


excel_path = os.path.join(OUTPUT_DIR, "churn_dashboard_report.xlsx")
wb.save(excel_path)
print(f"✅ Excel report saved → {excel_path}")
# ══════════════════════════════════════════════════════════
# STEP 5 — Export Power BI Dataset
# ══════════════════════════════════════════════════════════

print("\n📈 Preparing Power BI dataset...")

powerbi_df = df.copy()

# Customer Segment
powerbi_df["CustomerSegment"] = np.where(
    powerbi_df["MonthlyCharges"] >= powerbi_df["MonthlyCharges"].median(),
    "High Value",
    "Standard"
)

# Tenure Band
powerbi_df["TenureBand"] = pd.cut(
    powerbi_df["tenure"],
    bins=[0, 12, 24, 48, 72],
    labels=[
        "0-12 Months",
        "13-24 Months",
        "25-48 Months",
        "49-72 Months"
    ]
)

# Revenue Risk Flag
powerbi_df["RevenueRiskFlag"] = np.where(
    powerbi_df["Churn"] == "Yes",
    "At Risk",
    "Retained"
)

# Churn Category
powerbi_df["ChurnCategory"] = np.where(
    powerbi_df["Contract"] == "Month-to-month",
    "High Risk",
    "Low Risk"
)

powerbi_file = os.path.join(
    OUTPUT_DIR,
    "powerbi_dataset.csv"
)

powerbi_df.to_csv(
    powerbi_file,
    index=False
)

print(f"✅ Power BI dataset saved → {powerbi_file}")
# ══════════════════════════════════════════════════════════
# STEP 6 — Generate HTML Dashboard
# ══════════════════════════════════════════════════════════
print("\n Generating HTML dashboard...")

html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Customer Churn Analysis Dashboard</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family:'Segoe UI',sans-serif; background:#F5F5F5; color:#333; }}
  .header {{ background:linear-gradient(135deg,#1D9E75,#378ADD); color:white; padding:30px 40px; }}
  .header h1 {{ font-size:28px; margin-bottom:8px; }}
  .header p {{ font-size:14px; opacity:0.85; }}
  .kpi-grid {{ display:grid; grid-template-columns:repeat(4,1fr); gap:16px; padding:24px 40px; }}
  .kpi-card {{ background:white; border-radius:10px; padding:20px; text-align:center;
               box-shadow:0 2px 8px rgba(0,0,0,0.08); border-top:4px solid #378ADD; }}
  .kpi-value {{ font-size:28px; font-weight:700; color:#E24B4A; margin:8px 0; }}
  .kpi-label {{ font-size:12px; color:#888; text-transform:uppercase; letter-spacing:0.05em; }}
  .charts-grid {{ display:grid; grid-template-columns:repeat(2,1fr); gap:16px; padding:0 40px 24px; }}
  .chart-card {{ background:white; border-radius:10px; padding:20px;
                 box-shadow:0 2px 8px rgba(0,0,0,0.08); }}
  .chart-card h3 {{ font-size:14px; color:#555; margin-bottom:12px; font-weight:600; }}
  .chart-card img {{ width:100%; border-radius:6px; }}
  .footer {{ background:#1D9E75; color:white; text-align:center; padding:16px;
             font-size:13px; }}
</style>
</head>
<body>
<div class="header">
  <h1>Customer Churn Analysis Dashboard</h1>
  <p>Telco Customer Churn — End-to-End Analytics | Built by Sabari A S</p>
</div>
<div class="kpi-grid">
  <div class="kpi-card">
    <div class="kpi-label">Total Customers</div>
    <div class="kpi-value" style="color:#378ADD">{len(df):,}</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">Churned Customers</div>
    <div class="kpi-value">{int(df['Churn_Flag'].sum()):,}</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">Churn Rate</div>
    <div class="kpi-value">{df['Churn_Flag'].mean():.1%}</div>
  </div>
  <div class="kpi-card">
    <div class="kpi-label">Revenue at Risk</div>
    <div class="kpi-value">${df[df['Churn']=='Yes']['MonthlyCharges'].sum():,.0f}/mo</div>
  </div>
</div>
<div class="charts-grid">
  <div class="chart-card">
    <h3>Churn Distribution</h3>
    <img src="../visuals/01_churn_distribution.png">
  </div>
  <div class="chart-card">
    <h3>Churn by Contract Type</h3>
    <img src="../visuals/02_churn_by_contract.png">
  </div>
  <div class="chart-card">
    <h3>Churn by Internet Service</h3>
    <img src="../visuals/05_churn_by_internet.png">
  </div>
  <div class="chart-card">
    <h3>Churn by Payment Method</h3>
    <img src="../visuals/06_churn_by_payment.png">
  </div>
  <div class="chart-card">
    <h3>Tenure Distribution</h3>
    <img src="../visuals/03_tenure_distribution.png">
  </div>
  <div class="chart-card">
    <h3>Churn Heatmap</h3>
    <img src="../visuals/07_churn_heatmap.png">
  </div>
</div>
<div class="footer">
  Built by <strong>Sabari A S</strong> |
  <a href="https://linkedin.com/in/sabari3299" style="color:white">linkedin.com/in/sabari3299</a> |
  Open to Data Analyst roles
</div>
</body>
</html>"""
html_path = os.path.join(OUTPUT_DIR, "churn_dashboard.html")

with open(html_path, "w", encoding="utf-8") as f:
    f.write(html)

# ══════════════════════════════════════════════════════════
# STEP 7 — Final Summary
# ══════════════════════════════════════════════════════════
print("\n" + "="*60)
print("  AUTOMATION COMPLETE — ALL OUTPUTS GENERATED")
print("="*60)
print(f"\n📁 Output folder: {OUTPUT_DIR}")
print(f"\n   ✅ 8 PNG visuals        → {VISUALS_DIR}")
print(f"   ✅ Excel report         → churn_dashboard_report.xlsx")
print(f"   ✅ Power BI Dataset     → powerbi_dataset.csv")
print(f"   ✅ HTML dashboard       → churn_dashboard.html")
print(f"\n{'='*60}")
print("  Ready to present to any recruiter or stakeholder!")
print("="*60)